"""
output_writer.py
----------------
Genera el archivo SALIDAS_MONITOR.xlsx con cuatro hojas:
  - ESTADO_ACTUAL : snapshot completo de la corrida más reciente
  - CAMBIOS       : solo novedades — posiciones que cambiaron de estado
  - HISTORIAL     : log acumulativo de cada corrida
  - CONFIG        : tabla manual ticker / tipo (A o B) — no se toca
"""

import io
import logging
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLOR_VERDE    = "C6EFCE"
COLOR_AMARILLO = "FFEB9C"
COLOR_NARANJA  = "FFCC99"
COLOR_ROJO     = "FFC7CE"
COLOR_HEADER   = "2F4F8F"

EMOJI_COLOR = {
    "🟢": COLOR_VERDE,
    "🟡": COLOR_AMARILLO,
    "🟠": COLOR_NARANJA,
    "🔴": COLOR_ROJO,
}

# Orden de severidad para detectar si empeoró o mejoró
SEVERIDAD = {"🟢": 0, "🟡": 1, "🟠": 2, "🔴": 3}

COLS_ESTADO = [
    "fecha_run", "ticker", "categoria", "tipo_empresa", "tipo_ab",
    "cantidad", "coste_compra_usd", "valor_mercado_usd",
    "ganancia_pct", "ppp_equiv_usd", "fecha_primera_compra",
    "precio_actual_usd", "maximo_desde_entrada_usd",
    "sigma_mensual_12m",
    # Satélites
    "grupo_satelite", "rank_bruto", "rank_efectivo",
    "score_etf", "score_top5", "delta_momentum",
    "stop_s1_usd", "kxsigma",
    "s1_activado", "s2_decision", "s3_activado",
    # Tácticos
    "stop_t1_usd", "t1_activo", "t1_activado",
    "stop_t2_usd", "t2_activo", "t2_activado", "y_pct",
    "fase_modelo", "accion_modelo", "flag_revision", "tamano_posicion_pct",
    "t3_estado", "t3_precio_objetivo", "tir_objetivo",
    # Alerta final
    "alerta_emoji", "alerta_texto",
]

COLS_CAMBIOS = [
    "fecha_cambio", "ticker", "categoria",
    "alerta_anterior", "texto_anterior",
    "alerta_nueva", "texto_nuevo",
    "direccion",   # EMPEORÓ / MEJORÓ
]


def _apply_color(ws, row: int, col_start: int, col_end: int, hex_color: str):
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = fill


def _style_header(ws, n_cols: int):
    fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 55)


def construir_fila(run_ts: datetime, datos: dict) -> dict:
    row = {"fecha_run": run_ts}
    for col in COLS_ESTADO[1:]:
        row[col] = datos.get(col)
    return row


def _detectar_cambios(
    df_actual: pd.DataFrame,
    df_hist_prev: pd.DataFrame,
    run_ts: datetime,
) -> pd.DataFrame:
    """
    Compara el estado actual con la última corrida anterior por ticker.
    Retorna filas de cambios nuevos (solo cuando cambia el emoji de alerta).
    No repite un cambio si el ticker ya estaba en el mismo estado anterior.
    """
    if df_hist_prev.empty:
        return pd.DataFrame(columns=COLS_CAMBIOS)

    # Última corrida anterior por ticker
    df_prev = (
        df_hist_prev
        .sort_values("fecha_run")
        .groupby("ticker")
        .last()
        .reset_index()
    )[["ticker", "alerta_emoji", "alerta_texto"]]

    # Merge con estado actual
    df_comp = df_actual[["ticker", "categoria", "alerta_emoji", "alerta_texto"]].merge(
        df_prev, on="ticker", how="left", suffixes=("_nueva", "_anterior")
    )

    cambios = []
    for _, row in df_comp.iterrows():
        emoji_ant = str(row.get("alerta_emoji_anterior") or "🟢")
        emoji_nue = str(row.get("alerta_emoji_nueva") or "🟢")

        # Sin cambio → ignorar
        if emoji_ant == emoji_nue:
            continue

        sev_ant = SEVERIDAD.get(emoji_ant, 0)
        sev_nue = SEVERIDAD.get(emoji_nue, 0)
        direccion = "EMPEORÓ" if sev_nue > sev_ant else "MEJORÓ"

        cambios.append({
            "fecha_cambio":    run_ts,
            "ticker":          row["ticker"],
            "categoria":       row["categoria"],
            "alerta_anterior": emoji_ant,
            "texto_anterior":  row.get("alerta_texto_anterior", ""),
            "alerta_nueva":    emoji_nue,
            "texto_nuevo":     row.get("alerta_texto_nueva", ""),
            "direccion":       direccion,
        })

    return pd.DataFrame(cambios, columns=COLS_CAMBIOS)


def generar_excel(
    filas: list[dict],
    bytes_existente: bytes | None = None,
) -> bytes:
    run_ts = datetime.now()
    df_actual = pd.DataFrame([construir_fila(run_ts, f) for f in filas])
    df_actual = df_actual.reindex(columns=COLS_ESTADO)

    # ── Cargar hojas previas ───────────────────────────────────────────────
    df_hist_prev = pd.DataFrame(columns=COLS_ESTADO)
    df_cambios_prev = pd.DataFrame(columns=COLS_CAMBIOS)
    df_config = pd.DataFrame(columns=["ticker", "tipo", "tir_objetivo", "comentarios"])

    if bytes_existente:
        wb_prev = load_workbook(io.BytesIO(bytes_existente))
        for sheet in ["HISTORIAL", "CAMBIOS", "CONFIG"]:
            if sheet in wb_prev.sheetnames:
                try:
                    df_read = pd.read_excel(io.BytesIO(bytes_existente), sheet_name=sheet)
                    if sheet == "HISTORIAL":
                        df_hist_prev = df_read
                    elif sheet == "CAMBIOS":
                        df_cambios_prev = df_read
                    elif sheet == "CONFIG":
                        # Preservar TODAS las columnas que el usuario haya agregado.
                        # Solo actualizar si tiene datos — nunca pisar con vacío.
                        if not df_read.empty:
                            df_config = df_read
                            logger.info(
                                "CONFIG cargada: %d filas, columnas: %s",
                                len(df_config), list(df_config.columns)
                            )
                        else:
                            df_config = df_read  # preserva estructura aunque esté vacía
                            logger.warning("CONFIG existe pero está vacía — se preserva estructura.")
                except Exception as e:
                    logger.warning("No se pudo leer hoja %s: %s", sheet, e)
    else:
        logger.warning(
            "No hay Excel previo disponible — CONFIG se inicializa vacía. "
            "Si ya tenías datos en CONFIG, verificá que el archivo se descargó correctamente de Drive."
        )

    # ── Detectar cambios nuevos ────────────────────────────────────────────
    df_cambios_nuevos = _detectar_cambios(df_actual, df_hist_prev, run_ts)
    df_cambios = pd.concat([df_cambios_prev, df_cambios_nuevos], ignore_index=True)

    if not df_cambios_nuevos.empty:
        logger.info(
            "Cambios detectados: %d — %s",
            len(df_cambios_nuevos),
            df_cambios_nuevos[["ticker", "direccion", "alerta_nueva"]].to_dict("records"),
        )
    else:
        logger.info("Sin cambios de estado respecto a la corrida anterior.")

    # ── Historial acumulativo ──────────────────────────────────────────────
    df_hist = pd.concat([df_hist_prev, df_actual], ignore_index=True)

    # ── Escribir Excel ─────────────────────────────────────────────────────
    # Normalizar nombres de columnas de CONFIG (quitar espacios, lowercase → uppercase)
    df_config.columns = [str(c).strip().upper() for c in df_config.columns]
    # Asegurar que siempre existan las columnas mínimas esperadas
    for col in ["TICKER", "TIPO", "TIR_OBJETIVO", "COMENTARIOS"]:
        if col not in df_config.columns:
            df_config[col] = None

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_actual.to_excel(writer, sheet_name="ESTADO_ACTUAL", index=False)
        df_cambios.to_excel(writer, sheet_name="CAMBIOS", index=False)
        df_hist.to_excel(writer, sheet_name="HISTORIAL", index=False)
        df_config.to_excel(writer, sheet_name="CONFIG", index=False)

    buf.seek(0)
    wb = load_workbook(buf)

    # ── Estilo ESTADO_ACTUAL ───────────────────────────────────────────────
    ws = wb["ESTADO_ACTUAL"]
    _style_header(ws, len(COLS_ESTADO))
    emoji_col_idx = COLS_ESTADO.index("alerta_emoji") + 1
    for row_idx in range(2, ws.max_row + 1):
        emoji = str(ws.cell(row=row_idx, column=emoji_col_idx).value or "")
        color = EMOJI_COLOR.get(emoji, "FFFFFF")
        _apply_color(ws, row_idx, 1, len(COLS_ESTADO), color)
    _autofit(ws)
    ws.freeze_panes = "A2"

    # ── Estilo CAMBIOS ─────────────────────────────────────────────────────
    ws_c = wb["CAMBIOS"]
    _style_header(ws_c, len(COLS_CAMBIOS))
    # Colorear por alerta_nueva
    col_nueva_idx = COLS_CAMBIOS.index("alerta_nueva") + 1
    col_dir_idx   = COLS_CAMBIOS.index("direccion") + 1
    for row_idx in range(2, ws_c.max_row + 1):
        emoji = str(ws_c.cell(row=row_idx, column=col_nueva_idx).value or "")
        color = EMOJI_COLOR.get(emoji, "FFFFFF")
        _apply_color(ws_c, row_idx, 1, len(COLS_CAMBIOS), color)
        # Negrita en EMPEORÓ
        dir_cell = ws_c.cell(row=row_idx, column=col_dir_idx)
        if str(dir_cell.value) == "EMPEORÓ":
            dir_cell.font = Font(bold=True)
    _autofit(ws_c)
    ws_c.freeze_panes = "A2"

    # ── Estilo HISTORIAL ───────────────────────────────────────────────────
    ws_h = wb["HISTORIAL"]
    _style_header(ws_h, len(COLS_ESTADO))
    _autofit(ws_h)
    ws_h.freeze_panes = "A2"

    # ── Estilo CONFIG (no tocar datos, solo header) ────────────────────────
    ws_cfg = wb["CONFIG"]
    _style_header(ws_cfg, ws_cfg.max_column)
    _autofit(ws_cfg)

    out = io.BytesIO()
    wb.save(out)
    logger.info(
        "SALIDAS_MONITOR generado: %d posiciones | %d cambios nuevos | %d historial",
        len(df_actual), len(df_cambios_nuevos), len(df_hist),
    )
    return out.getvalue()
